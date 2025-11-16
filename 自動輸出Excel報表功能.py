import os
import glob
import pandas as pd
from openpyxl import load_workbook
import shutil
import json

def export_excel(raw_excel_path: str, txt_folder_path: str, data_postion_map: str, output_excel_path: str) -> bool:
    # =================================== 設定檔案路徑 ====================================
    # ... (註解) ...
    # ====================================== 設定區 ======================================
    try: # 【【新增】】
        shutil.copy(raw_excel_path, fr"{output_excel_path}\Product Test Report.xlsx")
        with open(data_postion_map, 'r', encoding='utf-8') as dict:
            test_data_map = json.load(dict)

        # ====================================== 副程式 ======================================
        """設定要使用的excel檔案"""
        def set_excel():
            excel_path = fr"{output_excel_path}\Product Test Report.xlsx"
            excel_workbook = load_workbook(excel_path)
            return (excel_path, excel_workbook)

        """取得指定路徑下的所有 txt 檔案"""
        def read_all_txt():
            txt_path = fr"{txt_folder_path}\*.txt"
            return glob.glob(txt_path)

        """取得所有工作表的名稱"""
        def read_all_sheetnames(excel_path):
            with pd.ExcelFile(excel_path) as excel:
                all_sheetnames = excel.sheet_names
            return all_sheetnames

        """取得 txt 檔案的純名稱"""
        def get_txt_name(txt_list):
            filename = os.path.basename(txt_list) # 去掉路徑
            filename = os.path.splitext(filename) # 去掉副檔名
            return filename[0]                    # 回傳 TXT 檔案名稱

        """選擇要開啟的 txt 檔案"""
        def read_and_clean_txt(txt_name):
            with open(txt_name, "r", encoding="utf-8") as document:        
                text_lines = [txt.strip() for txt in document.readlines()] # 去掉換行符號
            return text_lines

        """從 test_data_map 裡面讀取出 keyword 和 Test_data 填入 excel 的位置"""
        def read_step_keyword(sheetname):
            keyword = list(test_data_map[sheetname].keys())      # 讀取 dict 裡面的 keyword
            data_position = test_data_map[sheetname][keyword[0]] # 讀取 dict 裡面紀錄的 TXT 檔案資料位置
            return (keyword[0], data_position)                   # 回傳 dict 的內容

        """利用 keyword 找出需要的行"""
        def find_step_in_txt(keyword, text_lines):
            find_step = [txt for txt in text_lines if keyword in txt] # 在 TXT 檔案裡面找對應的 step 名稱
            if find_step:           # 如果有找到
                return find_step[0] # 回傳找到的 step 行
            else:                   # 或者什麼都不做
                return None         # 回傳空的內容

        """將找到的 txt 檔名填入 excel 檔案"""
        def write_filename_to_excel(excel_workbook, sheetname, filename, row):
            workbook = excel_workbook[sheetname] # 打開目前的工作表
            workbook[f"A{row}"] = filename       # 把目前的 TXT 檔案名稱填入
            return

        """將找到的 STEP 填入 excel 檔案"""
        def write_step_to_excel(excel_workbook, sheetname, find_step, col, row):
            workbook = excel_workbook[sheetname]
            if find_step is None:                # 沒找到 step 名稱
                workbook[f"{col}{row}"] = "fail" # 在 Excel 填入 fail
                return
            else:                                   # 找到 step 名稱
                workbook[f"{col}{row}"] = find_step # 在 Excel 填入 step 名稱
                return

        """找find_step的索引值"""
        def find_step_index(text_lines, find_step):
            if find_step is None: # 沒找到索引值
                return None       # 回傳 None
            else:
                step_index = text_lines.index(find_step) # 找到索引值
                return step_index                        # 回傳索引值

        """使用limit_list裡面抓出的數據來找limit填入excel"""
        def find_limit_and_write_to_excel(limit_position, step_index, text_lines, excel_workbook, sheetname, row):
            position = limit_position.split()
            row_offset = int(position[0])
            col_offset = int(position[1])
            excel_col = position[2]
            
            if step_index is None:
                workbook = excel_workbook[sheetname]
                workbook[f"{excel_col}{row}"] = "fail"
                return
            else:
                limit_line = text_lines[step_index + row_offset].split()
                limit = limit_line[col_offset - 1]
                workbook = excel_workbook[sheetname]
                workbook[f"{excel_col}{row}"] = limit
                return

        """所有資料填完後存檔"""
        def save_excel(excel_workbook, excel_path):
            excel_workbook.save(excel_path) # 存檔
            return

        # ====================================== 主程式 ======================================
        row = 2
        (excel_path, excel_workbook) = set_excel()       # 取得excel檔案和設定load_workbook 
        txt_list = read_all_txt()                        # 取得全部的txt檔案
        all_sheetnames = read_all_sheetnames(excel_path) # 取得excel裡所有的工作表名稱

        for n in range (len(txt_list)): # 這個迴圈負責遍歷所有txt檔案
            
            text_lines = read_and_clean_txt(txt_list[n]) # 把txt檔案的內容存進list裡面
            filename = get_txt_name(txt_list[n])         # 去掉txt檔案的路徑和副檔名，然後取得 TXT 檔案名稱
            
            for x in range(len(all_sheetnames)): # 這個迴圈負責跑一遍全部的工作表

                (keyword, data_position) = read_step_keyword(all_sheetnames[x])           # 讀出 dict 裡面的資訊
                find_step = find_step_in_txt(keyword, text_lines)                         # 用 dict 的 keyword 找到 TXT 檔案裡面的步驟名稱
                write_filename_to_excel(excel_workbook, all_sheetnames[x], filename, row) # 把目前開啟的 TXT 檔案名稱填入 Excel 報表
                step_index = find_step_index(text_lines, find_step)                       # 找出 step 名稱的 index
                write_step_to_excel(excel_workbook, all_sheetnames[x], find_step, data_position[0], row) # 把 step 名稱填入 Excel 報表

                for a in range (1, len(data_position)): # 跑一輪所需填入的 limit 資料
                    find_limit_and_write_to_excel(data_position[a], step_index, text_lines, excel_workbook, all_sheetnames[x], row) # 利用 dict 內容來找到所需的資料
            row += 1
        save_excel(excel_workbook, excel_path) # 迴圈結束後存檔
        
        # ==================================================================================
        return True # 【【修改】】

    except Exception as e: # 【【新增以下區塊】】
        print(f"❌ [P6] 執行 export_excel 時發生錯誤: {e}")
        return False