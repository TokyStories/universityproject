import pandas as pd
import numpy as np
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 步驟 1: 讀取並準備資料 ---
def create_report_template(input_csv_path: str, output_excel_path: str):
    # 定義你的 CSV 檔案路徑
    #csv_file_path = "stage2_output.csv"
    csv_file_path = input_csv_path
    # 讀取 CSV 檔案 
    try:
        df = pd.read_csv(csv_file_path)
    except FileNotFoundError:
        print(f"錯誤: 找不到 CSV 檔案 '{csv_file_path}'")
        return False # (修改)
    except Exception as e:
        print(f"讀取 CSV 時發生錯誤: {e}")
        return False # (修改)

    # (合併修改): 計算 Limits 數量並建立決策表 ---

    # (A) 核心邏輯:
    # 我們依 (parameter, Vin, step) 分組，並計算每個組的「原始行數」
    # "step 的數量"
    # .size() 會計算行數，.reset_index(name='limits_count') 將結果存為新欄
    config_df = df.groupby(['parameter', 'Vin', 'step'], dropna=False, sort=False).size().reset_index(name='limits_count')

    
    # 並且順序是正確的 (SHORT CKT TEST 會在第一個)

    # (B) 繼續建立我們原有的「決策輔助欄」
    # (1) 依 'parameter' 進行第一層分組
    # 【修改點 2】: 加入 sort=False 確保後續計算 (如 cumcount) 尊重此順序
    param_group = config_df.groupby('parameter', dropna=False, sort=False)

    # (2) 'param_size': (決策 A)
    config_df['param_size'] = param_group.transform('size')

    # (3) 'vin_nunique': (決策 B)
    config_df['vin_nunique'] = param_group['Vin'].transform('nunique')

    # (4) 'param_rank': (用於 Case 2a)
    config_df['param_rank'] = param_group.cumcount()

    # (5) 依 '(parameter, Vin)' 進行第二層分組 (用於決策 C)
    # 【修改點 3】: 加入 sort=False 確保後續計算 (如 cumcount) 尊重此順序
    sub_group = config_df.groupby(['parameter', 'Vin'], dropna=False, sort=False)

    # (6) 'sub_size': (決策 C)
    config_df['sub_size'] = sub_group.transform('size')

    # (7) 'sub_rank': (用於 Case 2b-ii)
    config_df['sub_rank'] = sub_group.cumcount()

    # --- 步驟 4: (新規則) 迭代並將結果存入變數 ---

    # 初始化一個空列表
    output_list = []

    # 我們迭代 config_df 來產生工作表名稱
    # 稍後我們也會迭代 config_df 來建立工作表
    for index, row in config_df.iterrows():
        
        # 取得所有需要的決策欄位
        param_size = row['param_size']
        vin_nunique = row['vin_nunique']
        sub_size = row['sub_size']
        parameter = row['parameter']
        vin = row['Vin']
        
        # 準備 'Vin' 字串
        vin_string = ""
        # 【修改】: 加上 str(vin).strip() 檢查，確保 vin 不是空字串
        if not pd.isna(vin) and str(vin).strip():
            vin_string = f" {int(float(vin))}V" # 【修改】: 同時改用 float() 再 int() 來處理可能的 ".0"
            
        # 執行三層決策樹
        if param_size == 1:
            output_list.append(f"{parameter}")
        else:
            if vin_nunique == 1:
                rank = int(row['param_rank'] + 1)
                output_list.append(f"{parameter}_{rank}")
            else:
                if sub_size == 1:
                    output_list.append(f"{parameter}{vin_string}")
                else:
                    rank = int(row['sub_rank'] + 1)
                    output_list.append(f"{parameter}{vin_string}_{rank}")

    # 驗證 'output_list' 變數的內容
    print("--- 步驟 4 完成: 已產生工作表名稱清單 ---")
    print(output_list)


    # --- 步驟 5 (全新): 建立 Excel 檔案、寫入動態標頭並套用樣式 ---

    # (A) 定義一個函式來清理工作表名稱
    def sanitize_sheet_name(name):
        # 移除 Excel 不允許的字元，包含我們發現的 '#'
        sanitized = name.replace('/', '_')
        sanitized = re.sub(r'\s*#\s*(\d+)', r'_\1', sanitized)
        sanitized = sanitized.replace('#', '_')
        invalid_chars_to_remove = r'[\[\]\*\\:\?]'
        sanitized = re.sub(invalid_chars_to_remove, '', sanitized)
        # 截斷到 Excel 允許的 31 個字元
        return sanitized[:31]

    # (B) 定義 Excel 標頭樣式 (字體、填滿、對齊)
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    # 淺灰色背景填滿 (您可以改成您範例中的顏色)
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # (C) 定義輸出的 Excel 檔案名稱
    excel_output_path = output_excel_path

    # (D) 使用 pd.ExcelWriter 來建立檔案並逐一寫入工作表
    try:
        with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
            
            print(f"--- 步驟 5 開始: 正在建立 Excel 檔案: {excel_output_path} ---")
            
            # 這次我們迭代 'config_df' 而不是 'output_list'
            # 因為我們需要 'limits_count' 欄位
            for index, row in config_df.iterrows():
                
                # (E) 取得工作表名稱 (使用與步驟 4 完全相同的邏輯)
                sheet_name_raw = output_list[index] # 我們可以直接使用 output_list 的結果
                sheet_name_clean = sanitize_sheet_name(sheet_name_raw)
                
                # (F) 準備此工作表的標頭 (Headers)
                static_headers = ['Product ID', 'Test Condition']
                
                # (G) 建立動態的 'Limits' 標頭
                limits_count = int(row['limits_count'])
                dynamic_headers = [f'Limits {i+1}' for i in range(limits_count)]
                
                # 組合所有標頭
                all_headers = static_headers + dynamic_headers
                
                # (H) 建立一個「只有標頭」的空 DataFrame
                sheet_df = pd.DataFrame(columns=all_headers)
                
                # (I) 將這個 DataFrame 寫入 Excel，這會自動建立標頭
                sheet_df.to_excel(writer, sheet_name=sheet_name_clean, index=False)
                
                # (J) 套用樣式：
                # 取得 'openpyxl' 的 worksheet 物件
                worksheet = writer.sheets[sheet_name_clean]
                
                # 迭代所有標頭儲存格 (A1, B1, C1...) 並套用樣式
                for col_idx in range(1, len(all_headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    cell = worksheet[f'{col_letter}1'] # A1, B1, C1...
                    
                    # 套用我們定義的樣式
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    
                    # (可選) 設定欄寬
                    worksheet.column_dimensions[col_letter].width = 20

                print(f"已建立工作表: '{sheet_name_clean}' (含 {limits_count} 個 Limits 欄位)")
                
        print("--- 步驟 5 完成: Excel 檔案建立完畢 ---")
        return True # (修改)

    except PermissionError:
        print(f"\n--- 錯誤 ---")
        print(f"權限錯誤: 無法寫入 '{excel_output_path}'。")
        print(f"請確認您已關閉該 Excel 檔案，並且您有權限寫入此資料夾。")
        return False # (修改)
    except Exception as e:
        print(f"\n--- 發生未預期的錯誤 ---")
        print(e)
        return False # (修改)